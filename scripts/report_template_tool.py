#!/usr/bin/env python3
import json
import sys
from copy import copy
from openpyxl import load_workbook


def clone_row_style(ws, src_row, dst_row, max_col):
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
        if src.number_format:
            dst.number_format = src.number_format


def ensure_rows(ws, template_row, start_row, needed_count, max_col):
    base_capacity = max(0, ws.max_row - start_row + 1)
    if needed_count <= base_capacity:
        return
    insert_at = ws.max_row + 1
    extra = needed_count - base_capacity
    ws.insert_rows(insert_at, extra)
    for offset in range(extra):
        clone_row_style(ws, template_row, insert_at + offset, max_col)


def clear_range(ws, start_row, end_row, cols):
    for row in range(start_row, end_row + 1):
        for col in cols:
            ws[f'{col}{row}'] = None


def fill_template(payload_path, template_path, output_path):
    with open(payload_path, 'r', encoding='utf-8') as f:
        payload = json.load(f)

    wb = load_workbook(template_path)
    ws = wb['Giriş']
    ws_completed = wb['Tamamlanan Frekanslar']
    ws_deviation = wb['Sapmalar']
    ws_groups = wb['Gruplar']

    params = payload.get('params', {})
    stats = payload.get('stats', {})
    visible_groups = payload.get('visibleGroups', [])[:5]
    all_groups = payload.get('allGroups', [])
    completed = payload.get('completedTasks', [])
    deviations = payload.get('deviationTasks', [])

    ws['G2'] = params.get('firma', '-')
    ws['G3'] = params.get('topLocation', '-')
    ws['G4'] = params.get('location', '-')
    ws['G5'] = params.get('reportDate', '-')
    ws['G6'] = params.get('reportDayCount', 1)
    ws['G7'] = params.get('requestedBy', '-')

    for idx, row_no in enumerate(range(14, 19)):
        row = visible_groups[idx] if idx < len(visible_groups) else None
        ws[f'B{row_no}'] = row.get('grup', '-') if row else '-'
        ws[f'E{row_no}'] = row.get('hedefFrekansSayisi', 0) if row else 0
        ws[f'H{row_no}'] = row.get('tamamlananFrekansSayisi', 0) if row else 0
        ws[f'K{row_no}'] = row.get('basariliIslemOrani', 0) / 100 if row else 0
        ws[f'O{row_no}'] = row.get('sapmaFrekansSayisi', 0) if row else 0
        ws[f'R{row_no}'] = row.get('kayipFrekansSayisi', 0) if row else 0
        ws[f'V{row_no}'] = row.get('genelOran', 0) / 100 if row else 0
        ws[f'AQ{row_no - 10}'] = row.get('grup', '-') if row else '-'
        ws[f'AV{row_no - 10}'] = row.get('hedefFrekansSayisi', 0) if row else 0
        ws[f'AZ{row_no - 10}'] = 0
        ws[f'BC{row_no - 10}'] = 0
        ws[f'BG{row_no - 10}'] = row.get('kayipFrekansSayisi', 0) if row else 0
        ws[f'BJ{row_no - 10}'] = 0
        ws[f'BM{row_no - 10}'] = 0

    ws['AK12'] = stats.get('totalFrequency', 0)
    ws['AK13'] = stats.get('completedFrequency', 0)
    ws['AK14'] = stats.get('realizedFrequency', 0)
    ws['AK15'] = stats.get('deviationFrequency', 0)
    ws['AK16'] = stats.get('lostFrequency', 0)
    ws['AK17'] = stats.get('successAverage', 0) / 100
    ws['AZ12'] = stats.get('totalFrequency', 0)
    ws['AZ13'] = stats.get('deviationFrequency', 0)
    ws['BN12'] = stats.get('totalFrequency', 0)
    ws['BN13'] = stats.get('lostFrequency', 0)

    ensure_rows(ws_completed, 4, 4, max(5, len(completed)), 7)
    clear_range(ws_completed, 4, ws_completed.max_row, ['A', 'B', 'C', 'D', 'E', 'F', 'G'])
    ws_completed['C3'] = stats.get('completedFrequency', 0)
    for idx, item in enumerate(completed, start=4):
        ws_completed[f'A{idx}'] = item.get('no', idx - 3)
        ws_completed[f'B{idx}'] = item.get('personel', '-')
        ws_completed[f'C{idx}'] = item.get('lokasyon', '-')
        ws_completed[f'D{idx}'] = item.get('gorevNo', '-')
        ws_completed[f'E{idx}'] = item.get('gorevTanimi', '-')
        ws_completed[f'F{idx}'] = item.get('tarihSaat', '-')
        ws_completed[f'G{idx}'] = item.get('durum', 'TAMAMLANDI')

    ensure_rows(ws_deviation, 4, 4, max(5, len(deviations)), 7)
    clear_range(ws_deviation, 4, ws_deviation.max_row, ['A', 'B', 'C', 'D', 'E', 'F', 'G'])
    ws_deviation['C3'] = stats.get('deviationFrequency', 0)
    for idx, item in enumerate(deviations, start=4):
        ws_deviation[f'A{idx}'] = item.get('no', idx - 3)
        ws_deviation[f'B{idx}'] = item.get('personel', '-')
        ws_deviation[f'C{idx}'] = item.get('lokasyon', '-')
        ws_deviation[f'D{idx}'] = item.get('gorevNo', '-')
        ws_deviation[f'E{idx}'] = item.get('gorevTanimi', '-')
        ws_deviation[f'F{idx}'] = item.get('tarihSaat', '-')
        ws_deviation[f'G{idx}'] = item.get('sapmaNedeni', 'Zamanında yapılamayan')

    ensure_rows(ws_groups, 3, 3, max(5, len(all_groups)), 10)
    clear_range(ws_groups, 3, ws_groups.max_row, ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'])
    ws_groups['D2'] = stats.get('totalFrequency', 0)
    ws_groups['E2'] = stats.get('totalFrequency', 0)
    ws_groups['F2'] = stats.get('completedFrequency', 0)
    ws_groups['G2'] = stats.get('deviationFrequency', 0)
    ws_groups['H2'] = stats.get('lostFrequency', 0)
    ws_groups['I2'] = stats.get('successAverage', 0) / 100
    ws_groups['J2'] = stats.get('completionRatio', 0) / 100
    for idx, item in enumerate(all_groups, start=3):
        ws_groups[f'A{idx}'] = item.get('no', idx - 2)
        ws_groups[f'B{idx}'] = item.get('grup', '-')
        ws_groups[f'C{idx}'] = item.get('lokasyon', '-')
        ws_groups[f'D{idx}'] = item.get('gunlukFrekansSayisi', 0)
        ws_groups[f'E{idx}'] = item.get('hedefFrekansSayisi', 0)
        ws_groups[f'F{idx}'] = item.get('tamamlananFrekansSayisi', 0)
        ws_groups[f'G{idx}'] = item.get('sapmaFrekansSayisi', 0)
        ws_groups[f'H{idx}'] = item.get('kayipFrekansSayisi', 0)
        ws_groups[f'I{idx}'] = item.get('basariliIslemOrani', 0) / 100
        ws_groups[f'J{idx}'] = item.get('genelOran', 0) / 100

    for coord in ['K14', 'K15', 'K16', 'K17', 'K18', 'V14', 'V15', 'V16', 'V17', 'V18', 'AK17']:
        ws[coord].number_format = '0.0%'
    for coord in ['I2', 'J2']:
        ws_groups[coord].number_format = '0.0%'
    wb.save(output_path)


if __name__ == '__main__':
    if len(sys.argv) != 4:
        raise SystemExit('usage: report_template_tool.py payload.json template.xlsx output.xlsx')
    fill_template(sys.argv[1], sys.argv[2], sys.argv[3])
