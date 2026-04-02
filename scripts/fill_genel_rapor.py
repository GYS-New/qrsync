#!/usr/bin/env python3
"""
QR-SYNC Genel Rapor Excel Doldurma Scripti
Şablonu koruyarak gerçek verilerle doldurur.
Grafikler, stiller ve merge yapısı korunur.
"""
import sys, json, copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

_thin_side = Side(border_style='thin')
_data_border_global = Border(top=_thin_side, bottom=_thin_side, left=_thin_side, right=_thin_side)
_no_border_global = Border()

# ─── Renk sabitleri ──────────────────────────────────────────────
GREEN_DARK  = '375623'
GREEN_HEAD  = 'D9E8D0'
YELLOW_FILL = 'FFC000'
WHITE       = 'FFFFFF'
GRAY_STRIPE = 'F4F8F4'

def write(ws, row, col, value, align='left'):
    merge_range = None
    for m in list(ws.merged_cells.ranges):
        if m.min_row == row and m.min_col == col:
            merge_range = str(m)
            break
    if merge_range:
        ws.unmerge_cells(merge_range)
    cell = ws.cell(row=row, column=col)
    cell.value = value
    try:
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    except Exception:
        pass
    if merge_range:
        ws.merge_cells(merge_range)

def copy_row_style(ws, src_row, dst_row, max_col=12):
    for col in range(1, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell.font      = copy.copy(src_cell.font)
            dst_cell.fill      = copy.copy(src_cell.fill)
            dst_cell.border    = copy.copy(src_cell.border)
            dst_cell.alignment = copy.copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
    if src_row in ws.row_dimensions:
        src_h = ws.row_dimensions[src_row].height
        if src_h:
            ws.row_dimensions[dst_row].height = src_h

def apply_cell_style(cell, bold=False, align='left', fill_color=None, border=True, font_color=None, font_size=11):
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    font_args = {'size': font_size, 'bold': bold}
    if font_color:
        font_args['color'] = font_color
    cell.font = Font(**font_args)
    if fill_color:
        cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)
    if border:
        cell.border = _data_border_global

def fill_table_sheet(ws, rows_data, col_writers, title_row=1, header_row=3, data_start=4):
    """
    Detay sheet'lerini (Tamamlanan, Sapmalar) doldurur.
    rows_data: list of dict
    col_writers: [(col_idx, key), ...]
    """
    TEMPLATE_END = data_start + 4  # şablondaki son örnek satır

    n = len(rows_data)
    if n > (TEMPLATE_END - data_start + 1):
        extra = n - (TEMPLATE_END - data_start + 1)
        ws.insert_rows(TEMPLATE_END + 1, extra)
        for i in range(extra):
            copy_row_style(ws, data_start, TEMPLATE_END + 1 + i)

    last_data_row = data_start + max(n - 1, TEMPLATE_END - data_start)
    max_col = max(c for c, _ in col_writers) + 1
    for r in range(data_start, last_data_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None

    # Toplam satırı güncelle (header_row + 1 = toplam satırı)
    total_row = header_row + 1
    ws.cell(row=total_row, column=3).value = n  # Toplam değeri

    for i, row_data in enumerate(rows_data):
        r = data_start + i
        ws.cell(row=r, column=1).value = row_data.get('sn', i + 1)
        for col_idx, key in col_writers:
            ws.cell(row=r, column=col_idx).value = row_data.get(key, '')

def fill_kayip_sheet(ws, rows_data):
    """
    Kayıp Frekanslar: SN(1), LOKASYON(2), GÖREV NO(3), GÖREV TANIMI(4), TARİH-SAAT(5), DURUM(6), KAYIP NEDENİ(7)
    """
    TEMPLATE_START = 4
    TEMPLATE_END   = 8
    n = len(rows_data)

    if n > (TEMPLATE_END - TEMPLATE_START + 1):
        extra = n - (TEMPLATE_END - TEMPLATE_START + 1)
        ws.insert_rows(TEMPLATE_END + 1, extra)
        for i in range(extra):
            copy_row_style(ws, TEMPLATE_START, TEMPLATE_END + 1 + i, max_col=7)

    last_data_row = TEMPLATE_START + max(n - 1, TEMPLATE_END - TEMPLATE_START)
    for r in range(TEMPLATE_START, last_data_row + 1):
        for col in range(1, 8):
            ws.cell(row=r, column=col).value = None

    # Toplam satırı
    ws.cell(row=3, column=3).value = n

    for i, row_data in enumerate(rows_data):
        r = TEMPLATE_START + i
        ws.cell(row=r, column=1).value = row_data.get('sn', i + 1)
        ws.cell(row=r, column=2).value = row_data.get('lokasyon', '')
        ws.cell(row=r, column=3).value = row_data.get('gorevNo', '')
        ws.cell(row=r, column=4).value = row_data.get('gorevTanimi', '')
        ws.cell(row=r, column=5).value = row_data.get('tarihSaat', '')
        ws.cell(row=r, column=6).value = row_data.get('durum', '')
        ws.cell(row=r, column=7).value = row_data.get('kayipNedeni', '')

def fill_frekans_disi_sheet(wb, rows_data):
    """
    Frekans Dışı sayfasını oluşturur veya doldurur.
    Columns: SN(A), ÜST LOKASYON(B), GRUP TANIMI(C), LOKASYON TANIMI(D), PERSONEL(E), TARİH-SAAT(F), AÇIKLAMA(G)
    """
    sheet_name = 'Frekans Dışı'
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Mevcut verileri temizle (başlık ve header korunur)
        for r in range(4, ws.max_row + 1):
            for c in range(1, 8):
                ws.cell(row=r, column=c).value = None
    else:
        ws = wb.create_sheet(sheet_name)

    col_widths = [6, 22, 28, 32, 24, 18, 40]
    col_letters = ['A','B','C','D','E','F','G']
    for i, w in enumerate(col_widths):
        ws.column_dimensions[col_letters[i]].width = w

    # Satır 1: Başlık
    ws.row_dimensions[1].height = 22
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = 'FREKANS DIŞI ÇALIŞMALAR'
    title_cell.font  = Font(bold=True, size=13, color=GREEN_DARK)
    title_cell.alignment = Alignment(horizontal='left', vertical='center')

    # Satır 2: Boş
    ws.row_dimensions[2].height = 8

    # Satır 3: Header
    headers = ['SN', 'ÜST LOKASYON', 'GRUP TANIMI', 'LOKASYON TANIMI', 'PERSONEL', 'TARİH-SAAT', 'AÇIKLAMA']
    ws.row_dimensions[3].height = 30
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=ci)
        cell.value = h
        apply_cell_style(cell, bold=True, align='center', fill_color=GREEN_DARK, font_color=WHITE, font_size=10)

    # Veri satırları
    n = len(rows_data)
    keys = ['sn','ustLokasyon','grupTanimi','lokasyonTanimi','personel','tarihSaat','aciklama']
    for i, row_data in enumerate(rows_data):
        r = 4 + i
        ws.row_dimensions[r].height = 18
        for ci, key in enumerate(keys, start=1):
            cell = ws.cell(row=r, column=ci)
            cell.value = row_data.get(key, '')
            cell.border = _data_border_global
            cell.alignment = Alignment(horizontal='center' if ci == 1 else 'left', vertical='center')
            cell.fill = PatternFill(fill_type='solid', fgColor=GRAY_STRIPE if i % 2 == 0 else WHITE)
            cell.font = Font(size=10)

    # Eğer veri yoksa bilgi satırı
    if n == 0:
        cell = ws.cell(row=4, column=1)
        cell.value = 'Bu rapor döneminde frekans dışı çalışma bulunmamaktadır.'
        cell.font = Font(italic=True, size=10, color='888888')
        cell.alignment = Alignment(horizontal='left', vertical='center')

def main():
    if len(sys.argv) != 4:
        print("Kullanim: fill_genel_rapor.py <template.xlsx> <payload.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    template_path, payload_path, output_path = sys.argv[1], sys.argv[2], sys.argv[3]

    with open(payload_path, 'r', encoding='utf-8') as f:
        d = json.load(f)

    wb = load_workbook(template_path)

    gruplar          = d.get('grupMetrikleri', [])
    toplam           = d.get('toplamGorev', 0)
    toplamTam        = d.get('toplamTamamlanan', 0)
    toplamSap        = d.get('toplamSapma', 0)
    toplamKay        = d.get('toplamKayip', 0)
    genelBasari      = d.get('genelBasari', 0)
    gunSayisi        = d.get('gunSayisi', 1)
    toplamGerceklesen = toplamTam + toplamSap

    toplamHedef = sum(g.get('hedef', 0) for g in gruplar)
    if toplamHedef == 0:
        toplamHedef = toplam

    # ══ GİRİŞ SAYFASI ══════════════════════════════════════════
    ws = wb['Giriş']

    # PARAMETRELER — sıra: Firma(2), Proje(3), Üst Lokasyon(4), Alt Lokasyon(5), RaporTarihi(6), GünSayısı(7), RaporuAlan(8)
    write(ws, 2, 7, d.get('firmaAdi', ''))
    write(ws, 3, 7, d.get('projeAdi', ''))
    write(ws, 4, 7, d.get('ustLokTanim', ''))
    write(ws, 5, 7, d.get('altLokTanim', ''))
    write(ws, 6, 7, d.get('raporTarihLabel', ''))
    write(ws, 7, 7, f"{gunSayisi} gün")
    write(ws, 8, 7, d.get('raporuAlan', 'Yönetim'))

    # HAKEDİŞ FAKTÖRLERİ — AQ(43)=Grup, AV(48)=Hedef, BG(59)=Kayıp
    for i, g in enumerate(gruplar[:20]):
        r = 4 + i
        write(ws, r, 43, g.get('grup', ''))
        write(ws, r, 48, g.get('hedef', 0), 'center')
        write(ws, r, 59, g.get('kayip', 0), 'center')

    # GRUP FREKANS GÖSTERGELERİ — satır 14'ten
    # B(2)=grup, E(5)=hedef, H(8)=tamamlanan, K(11)=basariOrani%, O(15)=sapma, R(18)=kayip, V(22)=genelOran
    n_grup = min(len(gruplar), 10)
    for i, g in enumerate(gruplar[:10]):
        r = 14 + i
        hedef_g = g.get('hedef', 0)
        tam_g   = g.get('tamamlanan', 0)
        sap_g   = g.get('sapma', 0)
        kayip_g = g.get('kayip', 0)
        basari_g = round(tam_g / hedef_g, 4) if hedef_g > 0 else 0
        genel_g  = round((tam_g + sap_g) / hedef_g, 4) if hedef_g > 0 else 0

        write(ws, r, 2,  g.get('grup', ''))
        write(ws, r, 5,  hedef_g, 'center')
        write(ws, r, 8,  tam_g,   'center')
        write(ws, r, 15, sap_g,   'center')

        kayip_cell = ws.cell(row=r, column=18)
        kayip_cell.value = kayip_g
        kayip_cell.alignment = Alignment(horizontal='center', vertical='center')

        cell_k = ws.cell(row=r, column=11)
        cell_k.value = basari_g
        cell_k.number_format = '0.00%'
        try:
            cell_k.alignment = Alignment(horizontal='center', vertical='center')
        except: pass

        cell_v = ws.cell(row=r, column=22)
        cell_v.value = genel_g
        cell_v.number_format = '0.00%'

    # Grafik referans güncelle
    last_grup_row = 13 + n_grup
    CHART_TITLES = {
        0: 'BAŞARILI İŞLEM ORANI',
        2: 'FREKANS SAPMALARI',
        3: 'GENEL DURUM',
        4: 'KAYIP FREKANS GÖSTERGELERİ',
    }

    def set_chart_title(chart, text):
        from openpyxl.drawing.text import CharacterProperties, ParagraphProperties
        try:
            for p in chart.title.tx.rich.p:
                if not p.pPr:
                    p.pPr = ParagraphProperties()
                if not p.pPr.defRPr:
                    p.pPr.defRPr = CharacterProperties()
                p.pPr.defRPr.sz = 1200
                p.pPr.defRPr.b = True
                if p.r:
                    p.r[0].t = text
                    if not p.r[0].rPr:
                        p.r[0].rPr = CharacterProperties()
                    p.r[0].rPr.sz = 1200
                    p.r[0].rPr.b = True
                    del p.r[1:]
                    return True
        except Exception:
            pass
        return False

    for idx, chart in enumerate(ws._charts):
        if idx == 0:
            try:
                chart.series[0].val.numRef.f = f'Giriş!$K$14:$K${last_grup_row}'
                chart.series[0].cat.strRef.f  = f'Giriş!$B$14:$B${last_grup_row}'
            except Exception:
                pass
        if idx in CHART_TITLES:
            set_chart_title(chart, CHART_TITLES[idx])
        try:
            for p2 in chart.title.spPr.txPr.p:
                if p2.r:
                    for run in p2.r:
                        if run.t == 'None':
                            run.t = CHART_TITLES.get(idx, '')
        except Exception:
            pass

    # FREKANS GÖSTERGELERİ — AA(27)/AK(37)
    toplamGerceklesen = toplamTam + toplamSap
    basari_pct = f'%{genelBasari}'
    frekans_vals = [toplamHedef, toplamTam, toplamGerceklesen, toplamSap, toplamKay, basari_pct]
    for i, val in enumerate(frekans_vals):
        write(ws, 12 + i, 37, val, 'center')

    # FREKANS SAPMALARI — AZ(52)
    sapma_orani = round(toplamSap / toplamHedef * 100) if toplamHedef > 0 else 0
    for i, val in enumerate([toplamHedef, toplamSap, f'%{sapma_orani}']):
        write(ws, 12 + i, 52, val, 'center')

    # KAYIP FREKANS GÖSTERGELERİ — BN(66)
    kayip_orani = round(toplamKay / toplamHedef * 100) if toplamHedef > 0 else 0
    for i, val in enumerate([toplamHedef, toplamKay, f'%{kayip_orani}']):
        write(ws, 12 + i, 66, val, 'center')

    # ══ TAMAMLANAN FREKANSLAR ══════════════════════════════════
    ws_tam = wb['Tamamlanan Frekanslar']
    rows_tam = d.get('tamamlananGorevler', [])
    fill_table_sheet(
        ws_tam, rows_tam,
        col_writers=[(2,'personel'),(3,'lokasyon'),(4,'gorevNo'),(5,'gorevTanimi'),(6,'tarihSaat'),(7,'durum')]
    )

    # ══ SAPMALAR ══════════════════════════════════════════════
    ws_sap = wb['Sapmalar']
    rows_sap = d.get('sapmaGorevler', [])
    fill_table_sheet(
        ws_sap, rows_sap,
        col_writers=[(2,'personel'),(3,'lokasyon'),(4,'gorevNo'),(5,'gorevTanimi'),(6,'tarihSaat'),(7,'sapmaNedeni')]
    )

    # ══ KAYIP FREKANSLAR ══════════════════════════════════════
    ws_kayip = wb['Kayıp Frekanslar']
    rows_kayip = d.get('kayipGorevler', [])
    fill_kayip_sheet(ws_kayip, rows_kayip)

    # ══ GRUPLAR ══════════════════════════════════════════════
    ws_gr = wb['Gruplar']
    n_gr = len(gruplar)

    # Kolon yapısı (GÖREV TANIMI kaldırıldı):
    # A(1)=SN, B(2)=GRUP, C(3)=LOKASYON, D(4)=GÜNLÜK FREKANS, E(5)=HEDEF,
    # F(6)=TAMAMLANAN, G(7)=SAPMA, H(8)=KAYIP, I(9)=BAŞARILI İŞLEM, J(10)=GENEL ORAN
    data_end = max(n_gr + 2, 3)
    ws_gr.cell(row=2, column=4).value  = f'=SUM(D3:D{data_end})'   # Günlük Frekans
    ws_gr.cell(row=2, column=5).value  = f'=SUM(E3:E{data_end})'   # Hedef
    ws_gr.cell(row=2, column=6).value  = f'=SUM(F3:F{data_end})'   # Tamamlanan
    ws_gr.cell(row=2, column=7).value  = f'=SUM(G3:G{data_end})'   # Sapma
    ws_gr.cell(row=2, column=8).value  = f'=SUM(H3:H{data_end})'   # Kayıp
    cell_i2 = ws_gr.cell(row=2, column=9)
    cell_i2.value = '=IF(E2>0,F2/E2,0)'
    cell_i2.number_format = '0%'
    cell_j2 = ws_gr.cell(row=2, column=10)
    cell_j2.value = '=IF(E2>0,(F2+G2)/E2,0)'
    cell_j2.number_format = '0%'

    clear_end = max(n_gr + 3, 10)
    for r in range(3, clear_end):
        for col in range(1, 11):
            ws_gr.cell(row=r, column=col).value = None
            ws_gr.cell(row=r, column=col).border = _no_border_global

    def pct_to_float(s):
        try: return float(str(s).replace('%','').replace(',','.') or 0) / 100.0
        except: return 0.0

    for i, g in enumerate(gruplar):
        r = 3 + i
        for col in range(1, 11):
            ws_gr.cell(row=r, column=col).border = _data_border_global
        hedef_val      = g.get('hedef', 0)
        tamamlanan_val = g.get('tamamlanan', 0)
        sapma_val      = g.get('sapma', 0)
        kayip_db       = g.get('kayip', 0)
        gunluk_fr      = g.get('gunlukFrekans', 0)

        ws_gr.cell(row=r, column=1).value = i + 1
        ws_gr.cell(row=r, column=2).value = g.get('grup', '')
        ws_gr.cell(row=r, column=3).value = g.get('lokasyon', '')
        ws_gr.cell(row=r, column=4).value = gunluk_fr
        ws_gr.cell(row=r, column=5).value = hedef_val
        ws_gr.cell(row=r, column=6).value = tamamlanan_val
        ws_gr.cell(row=r, column=7).value = sapma_val
        ws_gr.cell(row=r, column=8).value = kayip_db

        basari_val = round(tamamlanan_val / hedef_val, 4) if hedef_val > 0 else 0
        genel_val  = round((tamamlanan_val + sapma_val) / hedef_val, 4) if hedef_val > 0 else 0
        cell_i = ws_gr.cell(row=r, column=9)
        cell_i.value = basari_val
        cell_i.number_format = '0.00%'
        cell_j = ws_gr.cell(row=r, column=10)
        cell_j.value = genel_val
        cell_j.number_format = '0.00%'

    # ══ FREKANS DIŞI ══════════════════════════════════════════
    rows_fd = d.get('frekansDisiGorevler', [])
    fill_frekans_disi_sheet(wb, rows_fd)

    wb.save(output_path)
    print("OK")

if __name__ == '__main__':
    main()
