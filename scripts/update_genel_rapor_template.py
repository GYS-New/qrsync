#!/usr/bin/env python3
"""
QR-SYNC Genel Rapor Template Güncelleme Scripti
Tek seferlik çalıştırılır — template.xlsx yapısını yeni fill script'e uyarlar.

Değişiklikler:
  1. Giriş > Parametreler: Proje satırı eklenir (eskiden 7 satır → 8 satır)
  2. Kayıp Frekanslar: G sütunu header'ına "KAYIP NEDENİ" eklenir
  3. Gruplar: Header satırı 10 kolona güncellenir (GÖREV TANIMI kaldırılır)
  4. Frekans Dışı sheet: fill script zaten oluşturuyor, ön kontrol
"""
import sys, copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

GREEN_DARK = '375623'
WHITE = 'FFFFFF'
_thin = Side(border_style='thin')
_border = Border(top=_thin, bottom=_thin, left=_thin, right=_thin)

def apply_header_style(cell, text):
    cell.value = text
    cell.font = Font(bold=True, size=10, color=WHITE)
    cell.fill = PatternFill(fill_type='solid', fgColor=GREEN_DARK)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = _border

def main():
    if len(sys.argv) != 3:
        print("Kullanim: update_genel_rapor_template.py <input.xlsx> <output.xlsx>")
        sys.exit(1)

    src, dst = sys.argv[1], sys.argv[2]
    wb = load_workbook(src)

    # ══ 1. GİRİŞ — Proje satırı ekle ══════════════════════════════════
    ws_giris = wb['Giriş']

    # Mevcut row 3 değerini kontrol et — zaten Proje satırı varsa atla
    cur_row3 = ws_giris.cell(row=3, column=1).value
    if cur_row3 and 'proje' in str(cur_row3).lower():
        print("[Giriş] Proje satırı zaten mevcut, atlandı.")
    else:
        # Row 3'e insert et (mevcut 3-N, 4-N+1'e kayar)
        ws_giris.insert_rows(3)
        # Stil: üst satırdaki (row 2) etiket hücrelerini kopyala
        for col in range(1, 8):
            src_cell = ws_giris.cell(row=2, column=col)
            dst_cell = ws_giris.cell(row=3, column=col)
            if src_cell.has_style:
                dst_cell.font      = copy.copy(src_cell.font)
                dst_cell.fill      = copy.copy(src_cell.fill)
                dst_cell.border    = copy.copy(src_cell.border)
                dst_cell.alignment = copy.copy(src_cell.alignment)
        # Etiket: A sütununa "Proje" yaz (template'de genellikle A veya A:F merge'li)
        ws_giris.cell(row=3, column=1).value = 'Proje'
        ws_giris.cell(row=3, column=7).value = ''  # Değer fill script tarafından doldurulur
        print("[Giris] Proje satiri row 3'e eklendi.")

    # ══ 2. KAYIP FREKANSLAR — G kolonu header ══════════════════════════
    ws_kayip = wb['Kayıp Frekanslar']

    # Header satırını bul (genellikle row 3, G sütunu boşsa ekle)
    header_row = 3
    g_cell = ws_kayip.cell(row=header_row, column=7)
    if g_cell.value and 'kayip' in str(g_cell.value).lower().replace('ı','i').replace('ı','i'):
        print("[Kayıp Frekanslar] KAYIP NEDENİ header zaten mevcut, atlandı.")
    else:
        # F hücresinin stilini kopyalayarak G'ye KAYIP NEDENİ ekle
        f_cell = ws_kayip.cell(row=header_row, column=6)
        if f_cell.has_style:
            g_cell.font      = copy.copy(f_cell.font)
            g_cell.fill      = copy.copy(f_cell.fill)
            g_cell.border    = copy.copy(f_cell.border)
            g_cell.alignment = copy.copy(f_cell.alignment)
        apply_header_style(g_cell, 'KAYIP NEDENİ')
        # Sütun genişliği
        ws_kayip.column_dimensions['G'].width = 32
        print("[Kayıp Frekanslar] KAYIP NEDENİ header'ı G3'e eklendi.")

    # ══ 3. GRUPLAR — Header satırı 10 kolona güncelle ══════════════════
    ws_gr = wb['Gruplar']

    # Header satırı (genellikle row 1)
    # Yeni yapı: SN, GRUP, LOKASYON, GÜNLÜK FREKANS, HEDEF FREKANS,
    #            TAMAMLANAN FREKANS, SAPMA FREKANS, KAYIP FREKANS,
    #            BAŞARILI İŞLEM ORANI, GENEL ORAN
    new_headers = [
        'SN', 'GRUP', 'LOKASYON', 'GÜNLÜK FREKANS', 'HEDEF FREKANS',
        'TAMAMLANAN FREKANS', 'SAPMA FREKANS', 'KAYIP FREKANS',
        'BAŞARILI İŞLEM ORANI', 'GENEL ORAN'
    ]
    col_widths = [5, 22, 28, 14, 13, 16, 13, 13, 18, 13]

    # Header row bulma: row 1 kontrol et
    header_row_gr = 1
    for r in range(1, 4):
        v = ws_gr.cell(row=r, column=1).value
        if v and str(v).strip() in ('SN', 'Sn', '1'):
            header_row_gr = r
            break

    # Mevcut header'ı güncelle
    for ci, h in enumerate(new_headers, start=1):
        cell = ws_gr.cell(row=header_row_gr, column=ci)
        apply_header_style(cell, h)
        from openpyxl.utils import get_column_letter
        ws_gr.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]

    # Eğer 11. kolon (GÖREV TANIMI) varsa temizle
    cell_11 = ws_gr.cell(row=header_row_gr, column=11)
    if cell_11.value:
        for r in range(header_row_gr, ws_gr.max_row + 1):
            ws_gr.cell(row=r, column=11).value = None
        print("[Gruplar] 11. kolon (GÖREV TANIMI) temizlendi.")

    print("[Gruplar] Header satırı 10 kolona güncellendi.")

    # ══ 4. FREKANS DIŞI — fill script zaten oluşturuyor, bilgi ver ══════
    if 'Frekans Dışı' not in wb.sheetnames:
        print("[Frekans Dışı] Sheet yok — fill script çalışırken otomatik oluşturulacak.")
    else:
        print("[Frekans Dışı] Sheet zaten mevcut.")

    wb.save(dst)
    print(f"\nTemplate güncellendi: {dst}")

if __name__ == '__main__':
    main()
