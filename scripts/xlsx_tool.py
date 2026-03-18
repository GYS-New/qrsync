#!/usr/bin/env python3
import json, sys
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.chart import BarChart, LineChart, PieChart, Reference


def norm(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v)


def clean_color(color):
    if not color:
        return None
    rgb = getattr(color, 'rgb', None)
    if rgb:
        rgb = str(rgb)
        if rgb != '00000000':
            return rgb[-6:]
    return None


def border_css(side):
    style = getattr(side, 'style', None)
    color = clean_color(getattr(side, 'color', None)) or '000000'
    if not style:
        return None
    width = '1px'
    if style in ('medium', 'thick', 'double'):
        width = '2px'
    return f'{width} solid #{color}'


def extract_style(cell):
    fill = None
    if cell.fill and cell.fill.fill_type == 'solid':
        fill = clean_color(cell.fill.fgColor)
    font_color = clean_color(getattr(cell.font, 'color', None))
    return {
        'bg': f'#{fill}' if fill else None,
        'color': f'#{font_color}' if font_color else None,
        'bold': bool(getattr(cell.font, 'bold', False)),
        'italic': bool(getattr(cell.font, 'italic', False)),
        'fontSize': cell.font.sz,
        'hAlign': cell.alignment.horizontal,
        'vAlign': cell.alignment.vertical,
        'wrap': bool(getattr(cell.alignment, 'wrap_text', False)),
        'borderTop': border_css(cell.border.top),
        'borderRight': border_css(cell.border.right),
        'borderBottom': border_css(cell.border.bottom),
        'borderLeft': border_css(cell.border.left),
    }


def has_visible_style(style):
    return bool(
        style.get('bg') or style.get('color') or style.get('bold') or style.get('italic') or
        style.get('borderTop') or style.get('borderRight') or style.get('borderBottom') or style.get('borderLeft')
    )


def cmd_read(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(json.dumps({"headers": [], "rows": []}, ensure_ascii=False))
        return
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not r:
            continue
        item = {}
        empty = True
        for i, header in enumerate(headers):
            if not header:
                continue
            val = norm(r[i] if i < len(r) else None)
            if val is not None and val != '':
                empty = False
            item[header] = val
        if not empty:
            out.append(item)
    print(json.dumps({"headers": headers, "rows": out}, ensure_ascii=False))


def style_sheet(ws, headers, rows, widths=None):
    header_fill = PatternFill('solid', fgColor='2E8B2E')
    header_font = Font(color='FFFFFF', bold=True)
    wrap = Alignment(vertical='top', wrap_text=True)
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    for r_index, row in enumerate(rows, start=2):
        for c_index, val in enumerate(row, start=1):
            ws.cell(row=r_index, column=c_index, value=val).alignment = wrap
    for idx, h in enumerate(headers, start=1):
        width = widths[idx-1] if widths and idx-1 < len(widths) else None
        if not width:
            width = max(len(str(h or '')) + 4, 16)
            for row in rows[:200]:
                val = row[idx-1] if idx-1 < len(row) else ''
                width = min(max(width, len(str(val or '')) + 2), 50)
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.auto_filter.ref = ws.dimensions


def set_table_block(ws, start_row, headers, rows, widths=None):
    header_fill = PatternFill('solid', fgColor='2E8B2E')
    header_font = Font(color='FFFFFF', bold=True)
    wrap = Alignment(vertical='top', wrap_text=True)
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    for r_index, row_values in enumerate(rows, start=start_row + 1):
        for c_index, val in enumerate(row_values, start=1):
            ws.cell(row=r_index, column=c_index, value=val).alignment = wrap
    for idx, h in enumerate(headers, start=1):
        width = widths[idx-1] if widths and idx-1 < len(widths) else None
        if not width:
            width = max(len(str(h or '')) + 4, 16)
            for row_values in rows[:200]:
                val = row_values[idx-1] if idx-1 < len(row_values) else ''
                width = min(max(width, len(str(val or '')) + 2), 42)
        ws.column_dimensions[get_column_letter(idx)].width = width
    if headers:
        ws.auto_filter.ref = f'A{start_row}:{get_column_letter(len(headers))}{max(start_row, start_row + len(rows))}'


def make_chart(ws, excel_chart, header_map, start_row, row_count):
    chart_type = excel_chart.get('type', 'bar')
    category_key = excel_chart.get('categoryKey')
    series_keys = excel_chart.get('seriesKeys') or []
    if row_count <= 0 or not category_key or not series_keys or category_key not in header_map:
        return None
    if chart_type == 'pie':
        chart = PieChart()
        chart.height = excel_chart.get('heightRows', 14)
        chart.width = excel_chart.get('widthCells', 16)
        chart.style = excel_chart.get('style', 10)
        chart.title = excel_chart.get('title', '')
        chart.legend.position = 'b'
        chart.plotVisOnly = False
        value_col = header_map.get(series_keys[0]); category_col = header_map.get(category_key)
        if not value_col or not category_col: return None
        data = Reference(ws, min_col=value_col, min_row=start_row, max_row=start_row + row_count)
        cats = Reference(ws, min_col=category_col, min_row=start_row + 1, max_row=start_row + row_count)
        return chart, data, cats
    if chart_type == 'line':
        chart = LineChart(); chart.grouping = 'standard'; chart.marker = None; chart.y_axis.title = None
    else:
        chart = BarChart(); chart.type = 'col'; chart.style = excel_chart.get('style', 10); chart.gapWidth = 45 if chart_type == 'grouped_bar' else 30; chart.overlap = 0
    chart.title = excel_chart.get('title', '')
    chart.height = excel_chart.get('heightRows', 12); chart.width = excel_chart.get('widthCells', 18); chart.legend.position = 'r'; chart.plotVisOnly = False
    chart.y_axis.title = None; chart.x_axis.title = None; chart.x_axis.delete = False; chart.y_axis.delete = False
    first_series_col = header_map.get(series_keys[0]); last_series_col = header_map.get(series_keys[-1])
    if not first_series_col or not last_series_col: return None
    data = Reference(ws, min_col=first_series_col, max_col=last_series_col, min_row=start_row, max_row=start_row + row_count)
    cats = Reference(ws, min_col=header_map[category_key], min_row=start_row + 1, max_row=start_row + row_count)
    return chart, data, cats


def write_chart_sheet(ws, sheet):
    title = sheet.get('title') or sheet.get('name') or 'Grafik + Veri'
    subtitle = sheet.get('subtitle') or ''
    meta = sheet.get('meta') or []
    table = sheet.get('table') or {}
    excel_chart = sheet.get('excelChart') or {}
    ws.title = sheet.get('name', 'Grafik'); ws.sheet_view.showGridLines = False
    ws['A1'] = title; ws['A1'].font = Font(size=16, bold=True, color='163016'); ws.merge_cells('A1:H1')
    if subtitle:
        ws['A2'] = subtitle; ws['A2'].font = Font(size=11, italic=True, color='5E725E'); ws.merge_cells('A2:H2')
    row = 4
    if meta:
        ws[f'A{row}'] = 'Alan'; ws[f'B{row}'] = 'Değer'
        for cell in (ws[f'A{row}'], ws[f'B{row}']):
            cell.fill = PatternFill('solid', fgColor='2E8B2E'); cell.font = Font(color='FFFFFF', bold=True); cell.alignment = Alignment(vertical='top', wrap_text=True)
        for item in meta:
            row += 1
            ws[f'A{row}'] = item.get('label', ''); ws[f'B{row}'] = item.get('value', '')
            ws[f'A{row}'].font = Font(bold=True, color='284128'); ws[f'B{row}'].alignment = Alignment(vertical='top', wrap_text=True)
        row += 2
    headers = [h.get('label', h.get('key', '')) for h in table.get('headers', [])]
    keys = [h.get('key', '') for h in table.get('headers', [])]
    widths = [h.get('width') for h in table.get('headers', [])]
    rows = [item if isinstance(item, list) else [item.get(k, '') if item.get(k, '') is not None else '' for k in keys] for item in table.get('rows', [])]
    table_start_row = row + 20
    if headers: set_table_block(ws, table_start_row, headers, rows, widths)
    header_map = {key: idx + 1 for idx, key in enumerate(keys)}
    if headers and excel_chart:
        built = make_chart(ws, excel_chart, header_map, table_start_row, len(rows))
        if built:
            chart, data, cats = built
            chart.add_data(data, titles_from_data=True); chart.set_categories(cats); ws.add_chart(chart, excel_chart.get('anchor') or 'A4')


def infer_parameter_key(label):
    t = (label or '').strip().lower().replace(':', '').replace('ı', 'i').replace('İ', 'i')
    if 'firma' in t: return 'firmaId'
    if 'üst lokasyon' in t or 'ust lokasyon' in t: return 'ustLokasyonId'
    if 'alt lokasyon' in t: return 'altLokasyonId'
    if 'baslangic' in t or 'başlangıç' in t: return 'raporBaslangic'
    if 'bitis' in t or 'bitiş' in t: return 'raporBitis'
    if 'raporu alan' in t: return 'raporuAlan'
    if 'rapor tarihi' in t or 'raportarihi' in t or t == 'tarih': return 'raporTarihi'
    if 'rapor gün sayısı' in t or 'rapor gun sayisi' in t: return 'gunSayisi'
    return None


def _extract_range_values(value_ws, formula):
    if not formula or '!' not in formula:
        return []
    try:
        _, rng = formula.split('!', 1)
        rng = rng.replace('$', '')
        min_col, min_row, max_col, max_row = range_boundaries(rng)
    except Exception:
        return []
    values = []
    for r in range(min_row, max_row + 1):
        if min_col == max_col:
            values.append(norm(value_ws.cell(r, min_col).value) or '')
        else:
            parts = []
            for c in range(min_col, max_col + 1):
                val = norm(value_ws.cell(r, c).value)
                if val:
                    parts.append(val)
            values.append(' '.join(parts).strip())
    return values


def extract_chart(chart, value_ws):
    title = ''
    try:
        if chart.title and chart.title.tx and chart.title.tx.rich and chart.title.tx.rich.p:
            chunks = []
            for p in chart.title.tx.rich.p:
                if getattr(p, 'r', None):
                    for r in p.r:
                        if getattr(r, 't', None): chunks.append(r.t)
                elif getattr(p, 'endParaRPr', None) is not None:
                    pass
            title = ''.join(chunks).strip()
    except Exception:
        title = ''
    try:
        from_marker = chart.anchor._from
        to_marker = chart.anchor.to
        start_row = from_marker.row + 1; start_col = from_marker.col + 1
        end_row = to_marker.row + 1; end_col = to_marker.col + 1
    except Exception:
        start_row = start_col = 1; end_row = 12; end_col = 8
    kind = 'pie' if chart.__class__.__name__.lower().startswith('pie') else 'bar'
    categories = []
    series = []
    for idx, s in enumerate(getattr(chart, 'ser', []) or []):
        cat_formula = None
        if getattr(s, 'cat', None):
            cat_formula = getattr(getattr(s.cat, 'strRef', None), 'f', None) or getattr(getattr(s.cat, 'numRef', None), 'f', None)
        val_formula = getattr(getattr(getattr(s, 'val', None), 'numRef', None), 'f', None)
        cats = _extract_range_values(value_ws, cat_formula)
        vals = _extract_range_values(value_ws, val_formula)
        if cats and not categories:
            categories = cats
        values = []
        for v in vals:
            try:
                values.append(float(str(v).replace('%','').replace(',','.')))
            except Exception:
                values.append(0)
        name = ''
        try:
            tx_formula = getattr(getattr(s, 'tx', None), 'strRef', None)
            if tx_formula and getattr(tx_formula, 'f', None):
                extracted = _extract_range_values(value_ws, tx_formula.f)
                name = extracted[0] if extracted else ''
        except Exception:
            name = ''
        series.append({'name': name or f'Seri {idx+1}', 'values': values})
    return {
        'type': kind,
        'title': title,
        'from': {'row': start_row, 'col': start_col},
        'to': {'row': end_row, 'col': end_col},
        'categories': categories,
        'series': series,
    }


def build_sheet_model(style_ws, value_ws):
    max_row = min(max(style_ws.max_row, 1), 140)
    max_col = min(max(style_ws.max_column, 1), 90)
    merge_top_left = {}; merge_skip = set()
    for rng in style_ws.merged_cells.ranges:
        if rng.min_row > max_row or rng.min_col > max_col: continue
        rowspan = min(rng.max_row, max_row) - rng.min_row + 1; colspan = min(rng.max_col, max_col) - rng.min_col + 1
        merge_top_left[(rng.min_row, rng.min_col)] = (rowspan, colspan)
        for r in range(rng.min_row, min(rng.max_row, max_row) + 1):
            for c in range(rng.min_col, min(rng.max_col, max_col) + 1):
                if not (r == rng.min_row and c == rng.min_col): merge_skip.add((r, c))
    cells = []; inferred = []
    col_widths = {str(i): (style_ws.column_dimensions[get_column_letter(i)].width or 8.43) for i in range(1, max_col + 1)}
    row_heights = {str(i): (style_ws.row_dimensions[i].height or 15) for i in range(1, max_row + 1)}
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if (r, c) in merge_skip: continue
            scell = style_ws.cell(r, c); vcell = value_ws.cell(r, c); value = vcell.value
            if isinstance(value, str) and value.startswith('='): value = ''
            rowspan, colspan = merge_top_left.get((r, c), (1, 1))
            style = extract_style(scell)
            text_value = norm(value) or ''
            if not text_value and not has_visible_style(style):
                continue
            entry = {'r': r, 'c': c, 'rowspan': rowspan, 'colspan': colspan, 'value': text_value, 'style': style}
            cells.append(entry)
            key = infer_parameter_key(entry['value'])
            if key: inferred.append({'key': key, 'label': entry['value'], 'sheet': style_ws.title, 'cell': f'{get_column_letter(c)}{r}'})
    charts = [extract_chart(chart, value_ws) for chart in getattr(style_ws, '_charts', []) or []]
    return {
        'name': style_ws.title,
        'maxRow': max_row,
        'maxCol': max_col,
        'cells': cells,
        'colWidths': col_widths,
        'rowHeights': row_heights,
        'parameters': inferred,
        'showGridLines': bool(getattr(style_ws.sheet_view, 'showGridLines', True)),
        'orientation': getattr(style_ws.page_setup, 'orientation', None) or 'portrait',
        'charts': charts,
    }


def cmd_inspect_template(path):
    style_wb = load_workbook(path, data_only=False); value_wb = load_workbook(path, data_only=True)
    print(json.dumps({'sheets': [build_sheet_model(style_wb[name], value_wb[name]) for name in style_wb.sheetnames]}, ensure_ascii=False))


def cmd_write(payload_path, output_path):
    with open(payload_path, 'r', encoding='utf-8') as f: payload = json.load(f)
    wb = Workbook(); first = True
    for sheet in payload.get('sheets', []):
        ws = wb.active if first else wb.create_sheet(sheet.get('name', 'Sheet')); first = False
        if sheet.get('mode') == 'chart_with_table':
            write_chart_sheet(ws, sheet); continue
        ws.title = sheet.get('name', 'Sheet1')
        headers = [h.get('label', h.get('key', '')) for h in sheet.get('headers', [])]
        keys = [h.get('key', '') for h in sheet.get('headers', [])]
        widths = [h.get('width') for h in sheet.get('headers', [])]
        rows = [item if isinstance(item, list) else [item.get(k, '') if item.get(k, '') is not None else '' for k in keys] for item in sheet.get('rows', [])]
        style_sheet(ws, headers, rows, widths)
        for idx, row_values in enumerate(rows, start=2):
            for cidx, val in enumerate(row_values, start=1):
                ws.cell(row=idx, column=cidx, value=val)
    if not payload.get('sheets'): wb.active.title = 'Sheet1'
    wb.save(output_path); print(json.dumps({"ok": True, "output": output_path}, ensure_ascii=False))


def cmd_fill_template(template_path, payload_path, output_path):
    with open(payload_path, 'r', encoding='utf-8') as f: payload = json.load(f)
    wb = load_workbook(template_path)
    def get_top_left(ws, cell_ref):
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        try:
            col_letter, row = coordinate_from_string(cell_ref); col = column_index_from_string(col_letter)
        except Exception:
            return cell_ref
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
                return f'{get_column_letter(merged.min_col)}{merged.min_row}'
        return cell_ref
    for item in payload.get('cells', []):
        sheet_name = item.get('sheet'); cell_ref = item.get('cell'); value = item.get('value')
        if sheet_name not in wb.sheetnames: continue
        try: wb[sheet_name][get_top_left(wb[sheet_name], cell_ref)] = value
        except AttributeError: pass
    for table in payload.get('tables', []):
        sheet_name = table.get('sheet'); start_row = table.get('startRow', 1); start_col = table.get('startCol', 1)
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        for r_idx, row_values in enumerate(table.get('rows', [])):
            for c_idx, val in enumerate(row_values):
                target_row = start_row + r_idx; target_col = start_col + c_idx
                try: ws.cell(row=target_row, column=target_col).value = val
                except AttributeError:
                    ref = f'{get_column_letter(target_col)}{target_row}'
                    try: ws[get_top_left(ws, ref)] = val
                    except AttributeError: pass
    wb.save(output_path); print(json.dumps({"ok": True, "output": output_path}, ensure_ascii=False))


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('usage', file=sys.stderr); sys.exit(1)
    cmd = sys.argv[1]
    if cmd == 'read': cmd_read(sys.argv[2])
    elif cmd == 'write': cmd_write(sys.argv[2], sys.argv[3])
    elif cmd == 'fill_template': cmd_fill_template(sys.argv[2], sys.argv[3], sys.argv[4])
    elif cmd == 'inspect_template': cmd_inspect_template(sys.argv[2])
    else:
        print('unknown command', file=sys.stderr); sys.exit(1)
